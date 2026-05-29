"""
sync_engine.py — reusable HubSpot sync logic.
Auto-detects Excel column layout from header row, so it works across
different company files without hardcoding column numbers.
"""
import io
import time
import requests
import openpyxl
from typing import Optional, Callable

BASE_URL = "https://api.hubapi.com"

# HubSpot v3 association type IDs
ASSOC_CONTACT_COMPANY = 1
ASSOC_CONTACT_DEAL    = 4
ASSOC_COMPANY_DEAL    = 6

# Sheet name fragments that identify non-lead sheets to skip
NON_LEAD_KEYWORDS = ("funnel", "reporting", "leadgen", "basic", "pivot", "summary")


# ─── Column auto-detection ─────────────────────────────────────────────────

def _detect_columns(header_row: tuple, data_rows: list) -> dict:
    """
    Find 0-indexed column positions for the 6 fields we need by scanning
    the header row.  Also validates the 'name' column against real data rows
    to handle merged-cell offsets (e.g. the HVAC sheet quirk).
    """
    cols = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        val = str(cell).strip().lower()
        if val == "company":
            cols["company"] = i
        elif val == "name":
            cols["name"] = i
        elif val == "mail 1" and "email" not in cols:
            cols["email"] = i
        elif val == "contacted":
            cols["contacted"] = i
        elif val == "responded":
            cols["responded"] = i
        elif val == "meeting":
            cols["meeting"] = i

    # Validate 'name' against data (handles merged-cell offsets)
    if "name" in cols:
        idx = cols["name"]
        has_data = any(
            row and len(row) > idx and row[idx] is not None
            for row in data_rows[:10]
        )
        if not has_data:
            next_idx = idx + 1
            if any(row and len(row) > next_idx and row[next_idx] is not None
                   for row in data_rows[:10]):
                cols["name"] = next_idx

    return cols


def _as_bool(v) -> bool:
    return v is True or v == 1 or (isinstance(v, str) and v.strip().upper() == "TRUE")


# ─── Excel reader ──────────────────────────────────────────────────────────

def read_excel_bytes(file_bytes: bytes) -> tuple:
    """
    Read an Excel file from raw bytes.
    Returns:
        companies  – { company_name: [ contact_dict, … ] }
        sheets     – list of sheet names that were processed
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    companies: dict = {}
    sheets_processed: list = []

    for sheet_name in wb.sheetnames:
        # Skip non-lead sheets by name hint
        if any(kw in sheet_name.lower() for kw in NON_LEAD_KEYWORDS):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 4:
            continue  # not enough rows to have headers + data

        header_row = rows[2]      # row 3 (0-indexed 2) is always the column header
        data_rows  = rows[3:]     # data starts at row 4

        cols = _detect_columns(header_row, data_rows)

        required = {"company", "name", "email", "contacted", "responded", "meeting"}
        if not required.issubset(cols.keys()):
            continue  # sheet doesn't match the expected structure → skip

        max_col = max(cols.values())
        sheets_processed.append(sheet_name)

        for row in data_rows:
            if not row or len(row) <= max_col:
                continue

            company = row[cols["company"]]
            if not company:
                continue
            company = str(company).strip()
            if not company:
                continue

            # Email — skip "lusha" and other placeholders
            raw_email = row[cols["email"]]
            email: Optional[str] = None
            if raw_email:
                e = str(raw_email).strip()
                if e.lower() not in ("lusha", "none", ""):
                    email = e

            raw_name = row[cols["name"]]
            name = str(raw_name).strip().replace("\n", " ") if raw_name else ""

            companies.setdefault(company, []).append({
                "name":      name,
                "email":     email,
                "industry":  sheet_name,
                "contacted": _as_bool(row[cols["contacted"]]),
                "responded": _as_bool(row[cols["responded"]]),
                "meeting":   _as_bool(row[cols["meeting"]]),
            })

    wb.close()
    return companies, sheets_processed


def determine_stage(contacts: list) -> Optional[str]:
    """Return highest activity stage across all contacts, or None if no activity."""
    if any(c["meeting"]   for c in contacts): return "meeting"
    if any(c["responded"] for c in contacts): return "responded"
    if any(c["contacted"] for c in contacts): return "contacted"
    return None


def count_active(companies: dict) -> tuple:
    """Returns (total, active, contacts_with_email) counts."""
    active = {c: v for c, v in companies.items() if determine_stage(v) is not None}
    emails = sum(1 for contacts in active.values() for c in contacts if c["email"])
    return len(companies), len(active), emails


# ─── HubSpot API client ────────────────────────────────────────────────────

class HubSpot:
    def __init__(self, token: str):
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        })
        self._last_call = 0.0

    def _wait(self):
        gap = time.time() - self._last_call
        if gap < 0.13:
            time.sleep(0.13 - gap)
        self._last_call = time.time()

    def _get(self, path, **kw):
        self._wait()
        r = self.session.get(f"{BASE_URL}{path}", **kw)
        r.raise_for_status()
        return r.json()

    def _post(self, path, body):
        self._wait()
        r = self.session.post(f"{BASE_URL}{path}", json=body)
        r.raise_for_status()
        return r.json()

    def _patch(self, path, body):
        self._wait()
        r = self.session.patch(f"{BASE_URL}{path}", json=body)
        r.raise_for_status()
        return r.json()

    def _put(self, path):
        self._wait()
        r = self.session.put(f"{BASE_URL}{path}")
        r.raise_for_status()

    # ── Pipelines ──────────────────────────────────────────────────────────

    def get_pipelines(self) -> list:
        return self._get("/crm/v3/pipelines/deals").get("results", [])

    def get_pipeline_stages(self, pipeline_id: str) -> list:
        return self._get(f"/crm/v3/pipelines/deals/{pipeline_id}").get("stages", [])

    # ── Companies ──────────────────────────────────────────────────────────

    def find_company(self, name: str) -> Optional[str]:
        data = self._post("/crm/v3/objects/companies/search", {
            "filterGroups": [{"filters": [
                {"propertyName": "name", "operator": "EQ", "value": name}
            ]}],
            "properties": ["name"], "limit": 1,
        })
        res = data.get("results", [])
        return res[0]["id"] if res else None

    def create_company(self, name: str) -> str:
        return self._post("/crm/v3/objects/companies",
                          {"properties": {"name": name}})["id"]

    def get_or_create_company(self, name: str) -> tuple:
        cid = self.find_company(name)
        return (cid, False) if cid else (self.create_company(name), True)

    # ── Contacts ───────────────────────────────────────────────────────────

    def find_contact(self, email: str) -> Optional[str]:
        data = self._post("/crm/v3/objects/contacts/search", {
            "filterGroups": [{"filters": [
                {"propertyName": "email", "operator": "EQ", "value": email}
            ]}],
            "properties": ["email"], "limit": 1,
        })
        res = data.get("results", [])
        return res[0]["id"] if res else None

    def create_contact(self, email: str, first: str, last: str) -> str:
        return self._post("/crm/v3/objects/contacts", {
            "properties": {"email": email, "firstname": first, "lastname": last}
        })["id"]

    def get_or_create_contact(self, email: str, full_name: str) -> tuple:
        parts = (full_name or "").strip().split(None, 1)
        first = parts[0] if parts else ""
        last  = parts[1] if len(parts) > 1 else ""
        cid = self.find_contact(email)
        return (cid, False) if cid else (self.create_contact(email, first, last), True)

    # ── Deals ──────────────────────────────────────────────────────────────

    def get_company_deals(self, company_id: str) -> list:
        data = self._get(f"/crm/v3/objects/companies/{company_id}/associations/deals")
        return [x["id"] for x in data.get("results", [])]

    def create_deal(self, name: str, pipeline_id: str, stage_id: str) -> str:
        return self._post("/crm/v3/objects/deals", {
            "properties": {
                "dealname": name,
                "pipeline": pipeline_id,
                "dealstage": stage_id,
            }
        })["id"]

    def update_deal_stage(self, deal_id: str, stage_id: str) -> None:
        self._patch(f"/crm/v3/objects/deals/{deal_id}",
                    {"properties": {"dealstage": stage_id}})

    # ── Associations ───────────────────────────────────────────────────────

    def _associate(self, from_t, from_id, to_t, to_id, type_id):
        try:
            self._put(f"/crm/v3/objects/{from_t}/{from_id}"
                      f"/associations/{to_t}/{to_id}/{type_id}")
        except requests.HTTPError as e:
            if e.response is not None and e.response.status_code == 409:
                return  # already associated — fine
            raise

    def link_contact_company(self, ct_id, co_id):
        self._associate("contacts", ct_id, "companies", co_id, ASSOC_CONTACT_COMPANY)

    def link_contact_deal(self, ct_id, d_id):
        self._associate("contacts", ct_id, "deals", d_id, ASSOC_CONTACT_DEAL)

    def link_company_deal(self, co_id, d_id):
        self._associate("companies", co_id, "deals", d_id, ASSOC_COMPANY_DEAL)


# ─── Main sync logic ───────────────────────────────────────────────────────

def run_sync(companies: dict, hs: HubSpot,
             pipeline_id: str, stage_map: dict,
             log: Callable = print) -> dict:
    """
    Sync all active companies to HubSpot.
    `log` is called with each status message for real-time UI updates.
    Returns a stats dict.
    """
    stats = {
        "skipped": 0, "co_created": 0, "ct_created": 0,
        "deal_created": 0, "deal_updated": 0, "errors": 0,
    }

    for company_name, contacts in companies.items():
        stage = determine_stage(contacts)
        if stage is None:
            stats["skipped"] += 1
            continue

        log(f"▸  {company_name}  →  {stage.upper()}")

        # Company
        try:
            co_id, created = hs.get_or_create_company(company_name)
            if created:
                stats["co_created"] += 1
                log(f"   + company created")
        except requests.HTTPError as e:
            code = e.response.status_code if e.response else "?"
            log(f"   ! company error HTTP {code} — skipping")
            stats["errors"] += 1
            continue

        # Contacts
        ct_ids = []
        for c in contacts:
            if not c["email"]:
                continue
            try:
                ct_id, created = hs.get_or_create_contact(c["email"], c["name"])
                if created:
                    stats["ct_created"] += 1
                    log(f"   + contact  {c['name']} <{c['email']}>")
                ct_ids.append(ct_id)
                hs.link_contact_company(ct_id, co_id)
            except requests.HTTPError as e:
                code = e.response.status_code if e.response else "?"
                log(f"   ! contact error {c['email']}  HTTP {code}")
                stats["errors"] += 1

        # Deal
        stage_id = stage_map[stage]
        try:
            existing = hs.get_company_deals(co_id)
        except Exception:
            existing = []

        try:
            if existing:
                deal_id = existing[0]
                hs.update_deal_stage(deal_id, stage_id)
                stats["deal_updated"] += 1
                log(f"   = deal stage → {stage}")
            else:
                deal_id = hs.create_deal(company_name, pipeline_id, stage_id)
                stats["deal_created"] += 1
                log(f"   + deal created  (stage={stage})")

            hs.link_company_deal(co_id, deal_id)
            for ct_id in ct_ids:
                hs.link_contact_deal(ct_id, deal_id)

        except requests.HTTPError as e:
            code = e.response.status_code if e.response else "?"
            log(f"   ! deal error  HTTP {code}")
            stats["errors"] += 1

    return stats
